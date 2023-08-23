import openpyxl
import os
import docx

wb = openpyxl.load_workbook('Акт сверки.xlsx')
ws = wb.active
NowRow = 10
ActVal = 1
Date = ""

if __name__ == "__main__":
    os.chdir('/CargoTransportation/Акт Сверки')
    with open('input_file.txt', encoding="utf-8") as input_file:
        input_file = input_file.read()

    input_file = [[i.split('\t') for i in day.split('\n')] for day in input_file.split('\n\n')]
    for day in input_file:
        count = 0
        for way in day:
            count += int(way[3])
            if way[0] != "":
                ws[f"B{NowRow}"] = way[0]
                Date = way[0]
        ws[f"C{NowRow}"] = f"Приход ({ActVal} от {Date})"
        ws[f"D{NowRow}"] = str(count)
        NowRow += 1
        ActVal += 1

os.chdir('C:\\Users\\Ryzhk\\Desktop')
wb.save("Акт сверки Уралвзрывпром.xlsx")
